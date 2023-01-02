import time
from datetime import datetime
print(time.ctime())
st = time.time()
import arcpy
import pandas
import settings
import exl_reader
import os
from openpyxl import load_workbook 
import shutil
from xlwt import Workbook
from openpyxl.worksheet.datavalidation import DataValidation


print("========== Step 5 ==========")
#read the cursor fields from the config file and write to Excel file template - "Report"
file = "C:\DEV\MIGRASH2\Report.xlsx"
shutil.copyfile('C:\DEV\connectionfile\Report_Temp.xlsx', file)
xfile = openpyxl.load_workbook(file)
for layer in exl_reader.JsonOutput:
      if layer != "migrashim_humim": 
            print ("Now making a report to {} -  Layer {} from {}".format(layer,list(exl_reader.JsonOutput).index(layer) + 1,len(exl_reader.JsonOutput)))
            items = exl_reader.JsonOutput.get(layer)
            layer_fields = settings.Convert(items["ATTRIBUTES"])              
            with arcpy.da.SearchCursor(str(GDB) +'\\'+ layer + "_Select", layer_fields) as cur:
                  df = pandas.DataFrame((cur), columns = layer_fields)
                  sheet = xfile.get_sheet_by_name('{}'.format(items["HEB_NAME"]))
                  for col in layer_fields:
                        sheet.cell(row = 1 ,column =layer_fields.index(col) +1).value = col
                        for ro in range(1, len(df) + 1):
                              sheet.cell(row = ro + 1 ,column =layer_fields.index(col) +1).value = df[col][ro-1]

dv = DataValidation(type="list", formula1='A1:A10000', allow_blank=True)
sheet = xfile.get_sheet_by_name("סימולטור")
dv.add(sheet["C1"])
sheet.add_data_validation(dv)
xfile.save(file)


#======================================================================================================================================
print("========== Step 6 ==========")

et = time.time()
res = et - st
final_res = str(res / 60)
print("Done!")
print("Good work {}! ".format(settings.username))
print("\N{smiling face with sunglasses}")
print(time.ctime())
print('Execution time:', final_res[0:final_res.find(".") + 3], 'minutes')

                  









#=======================================================================================================================================
print("========== Step 3 ==========")
#Combines the layers to "yeudei_karka" layer by Spatial join and measures whether there is an overlap or closeness of up to 100 meters from the layer by "Distance" column.

for layer in exl_reader.JsonOutput:
      if layer != "migrashim_humim":
            items = exl_reader.JsonOutput.get(layer)
            print ("Now making a spatial join to  {} -  Layer {} from {}".format(layer,list(exl_reader.JsonOutput).index(layer) + 1,len(exl_reader.JsonOutput)))
            arcpy.SpatialJoin_analysis(str(GDB) +'\\'+ layer, str(GDB) +"\migrashim_humim_For_Join",str(GDB) +'\\'+ layer + "_join","JOIN_ONE_TO_MANY", "KEEP_ALL",None,"CLOSEST", "{} Meters".format(items["BUFFER"]), "Distance")
      

print("Step 3 completed successfully!")

#=======================================================================================================================================
print("========== Step 4 ==========")
#Creates a new layer that shows only the values whose distance is between 0 - 100 by Select tool
for layer in exl_reader.JsonOutput:
      if layer != "migrashim_humim":
            print ("now filtering {} -  Layer {} from {}".format(layer,list(exl_reader.JsonOutput).index(layer) + 1,len(exl_reader.JsonOutput)))
            arcpy.analysis.Select(str(GDB) +'\\'+ layer + "_join", str(GDB) +'\\'+ layer + "_Select", "Distance <> -1")

print("Step 4 completed successfully!")

#=======================================================================================================================================

